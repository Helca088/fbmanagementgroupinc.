from io import BytesIO
from datetime import datetime
from django.http import HttpResponse
from django.utils import timezone
from django.db.models import Count, Q, F

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER
from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer,
)

from tickets.models import (
    Ticket,
    Technician,
    TicketStatusLog,
    TicketAssignmentLog,
    TicketAdditionalAssignmentLog,
    Outlet,
)

from tickets.views import get_filtered_tickets


def get_export_tickets(request):

    tickets = Ticket.objects.all()

    start = request.GET.get("start", "").strip()
    end = request.GET.get("end", "").strip()
    department = request.GET.get("department", "").strip()
    outlet = request.GET.get("outlet", "").strip()

    if start and end:
        tickets = tickets.filter(
            created_at__date__range=[start, end]
        )

    elif start:
        tickets = tickets.filter(
            created_at__date__gte=start
        )

    elif end:
        tickets = tickets.filter(
            created_at__date__lte=end
        )

    if department:
        tickets = tickets.filter(
            department__name=department
        )

    if outlet:
        tickets = tickets.filter(
            outlet_id=outlet
        )

    return tickets

def style_excel_sheet(ws, title, headers):

    ws.merge_cells(
        start_row=1,
        start_column=1,
        end_row=1,
        end_column=len(headers)
    )

    title_cell = ws.cell(row=1, column=1)
    title_cell.value = title
    title_cell.font = Font(
        bold=True,
        size=16
    )
    title_cell.alignment = Alignment(
        horizontal="center"
    )

    for col, header in enumerate(headers, start=1):

        cell = ws.cell(
            row=3,
            column=col
        )

        cell.value = header

        cell.font = Font(
            bold=True,
            color="FFFFFF"
        )

        cell.fill = PatternFill(
            "solid",
            fgColor="2563EB"
        )

        cell.alignment = Alignment(
            horizontal="center"
        )


def auto_width(ws):

    for column_cells in ws.columns:

        length = 0

        column = column_cells[0].column

        for cell in column_cells:

            if cell.value:

                length = max(
                    length,
                    len(str(cell.value))
                )

        ws.column_dimensions[
            get_column_letter(column)
        ].width = min(
            length + 3,
            50
        )

def export_tickets_excel(request):

    tickets = get_export_tickets(request)

    wb = Workbook()

    ws = wb.active
    ws.title = "Tickets"

    headers = [
        "Outlet",
        "Created",
        "Department",
        "Concern Type",
        "Message",
    ]

    style_excel_sheet(
        ws,
        "FB Management - Ticket Report",
        headers
    )

    row = 4

    for ticket in tickets.order_by("-created_at"):

        ws.cell(
            row=row,
            column=1,
            value=(
                ticket.outlet.name
                if ticket.outlet
                else ""
            )
        )

        ws.cell(
            row=row,
            column=2,
            value=(
                ticket.created_at.strftime(
                    "%Y-%m-%d %H:%M"
                )
                if ticket.created_at
                else ""
            )
        )

        ws.cell(
            row=row,
            column=3,
            value=(
                ticket.department.name
                if ticket.department
                else ""
            )
        )

        ws.cell(
            row=row,
            column=4,
            value=(
                ticket.concern_type.name
                if ticket.concern_type
                else ""
            )
        )

        ws.cell(
            row=row,
            column=5,
            value=ticket.message or ""
        )

        row += 1

    auto_width(ws)

    output = BytesIO()

    wb.save(output)

    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    )

    response["Content-Disposition"] = (
        'attachment; filename="tickets_report.xlsx"'
    )

    return response

def export_tickets_pdf(request):

    tickets = get_export_tickets(request)

    output = BytesIO()

    document = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        rightMargin=25,
        leftMargin=25,
        topMargin=25,
        bottomMargin=25,
    )

    styles = getSampleStyleSheet()

    # ============================================================
    # STYLES
    # ============================================================

    title_style = ParagraphStyle(
        "ReportTitle",
        parent=styles["Title"],
        fontSize=18,
        alignment=TA_CENTER,
        spaceAfter=15,
    )

    # Message style
    # This allows long messages to wrap inside the table cell
    message_style = ParagraphStyle(
        "MessageStyle",
        parent=styles["Normal"],
        fontSize=8,
        leading=10,
        wordWrap="CJK",
    )

    elements = []

    # ============================================================
    # TITLE
    # ============================================================

    elements.append(
        Paragraph(
            "FB Management — Ticket Report",
            title_style
        )
    )

    elements.append(
        Spacer(1, 10)
    )

    # ============================================================
    # TABLE DATA
    # ============================================================

    data = [[
        "Outlet",
        "Created",
        "Department",
        "Concern Type",
        "Message",
    ]]

    for ticket in tickets.order_by("-created_at"):

        data.append([
            ticket.outlet.name
            if ticket.outlet else "",

            ticket.created_at.strftime(
                "%Y-%m-%d %H:%M"
            )
            if ticket.created_at else "",

            ticket.department.name
            if ticket.department else "",

            ticket.concern_type.name
            if ticket.concern_type else "",

            # IMPORTANT:
            # Use Paragraph here so long messages wrap
            # instead of overflowing outside the table.
            Paragraph(
                str(ticket.message or ""),
                message_style
            ),
        ])

    # ============================================================
    # TABLE
    # ============================================================

    table = Table(
        data,
        repeatRows=1,
        colWidths=[
            100,  # Outlet
            100,  # Created
            100,  # Department
            130,  # Concern Type
            300,  # Message
        ]
    )

    # ============================================================
    # TABLE STYLE
    # ============================================================

    table.setStyle(
        TableStyle([

            # Header background
            (
                "BACKGROUND",
                (0, 0),
                (-1, 0),
                colors.HexColor("#2563EB")
            ),

            # Header text
            (
                "TEXTCOLOR",
                (0, 0),
                (-1, 0),
                colors.white
            ),

            # Header font
            (
                "FONTNAME",
                (0, 0),
                (-1, 0),
                "Helvetica-Bold"
            ),

            # Grid
            (
                "GRID",
                (0, 0),
                (-1, -1),
                0.5,
                colors.HexColor("#E5E7EB")
            ),

            # Vertical alignment
            (
                "VALIGN",
                (0, 0),
                (-1, -1),
                "TOP"
            ),

            # Font size
            (
                "FONTSIZE",
                (0, 0),
                (-1, -1),
                8
            ),

            # Cell padding
            (
                "LEFTPADDING",
                (0, 0),
                (-1, -1),
                6
            ),

            (
                "RIGHTPADDING",
                (0, 0),
                (-1, -1),
                6
            ),

            (
                "TOPPADDING",
                (0, 0),
                (-1, -1),
                5
            ),

            (
                "BOTTOMPADDING",
                (0, 0),
                (-1, -1),
                5
            ),

            # Alternating row colors
            (
                "ROWBACKGROUNDS",
                (0, 1),
                (-1, -1),
                [
                    colors.white,
                    colors.HexColor("#F8FAFC")
                ]
            ),
        ])
    )

    elements.append(table)

    # ============================================================
    # BUILD PDF
    # ============================================================

    document.build(elements)

    output.seek(0)

    # ============================================================
    # RESPONSE
    # ============================================================

    response = HttpResponse(
        output.getvalue(),
        content_type="application/pdf"
    )

    response["Content-Disposition"] = (
        'attachment; filename="tickets_report.pdf"'
    )

    return response

# ============================================================
# OUTLET EXPORT — PDF
# ============================================================

def export_outlets_pdf(request):

    # Use EXACT SAME FILTERS as Outlet report page
    tickets = get_filtered_tickets(request)

    # ========================================================
    # OUTLET SUMMARY
    # ========================================================

    outlet_summary = (
        tickets
        .values("outlet__name")
        .annotate(
            total=Count("id"),

            pending=Count(
                "id",
                filter=Q(status="pending")
            ),

            progress=Count(
                "id",
                filter=Q(status="progress")
            ),

            resolved=Count(
                "id",
                filter=Q(status="resolved")
            ),

            cancelled=Count(
                "id",
                filter=Q(status="cancelled")
            ),
        )
        .order_by("outlet__name")
    )

    # ========================================================
    # CONCERNS PER OUTLET
    # ========================================================

    concerns_per_outlet = (
        tickets
        .values(
            "outlet__name",
            "concern_type__name"
        )
        .annotate(
            total=Count("id")
        )
        .order_by(
            "outlet__name",
            "-total"
        )
    )

    # ========================================================
    # RESPONSE
    # ========================================================

    response = HttpResponse(
        content_type="application/pdf"
    )

    selected_outlet = request.GET.get("outlet", "").strip()
    selected_department = request.GET.get("department", "").strip()

    if selected_outlet:
        outlet_obj = Outlet.objects.filter(
            id=selected_outlet
        ).first()

        outlet_name = (
            outlet_obj.name
            if outlet_obj
            else "Outlet"
        )

        filename = (
            f"{outlet_name}_outlet_report.pdf"
            .replace(" ", "_")
        )

    elif selected_department:

        filename = (
            f"{selected_department}_outlet_report.pdf"
            .replace(" ", "_")
        )

    else:

        filename = "outlets_report.pdf"

    response["Content-Disposition"] = (
        f'attachment; filename="{filename}"'
    )

    # ========================================================
    # PDF DOCUMENT
    # ========================================================

    from reportlab.lib.pagesizes import landscape, A4

    doc = SimpleDocTemplate(
        response,
        pagesize=landscape(A4),
        rightMargin=30,
        leftMargin=30,
        topMargin=30,
        bottomMargin=30,
    )

    styles = getSampleStyleSheet()

    elements = []

    # ========================================================
    # TITLE
    # ========================================================

    elements.append(
        Paragraph(
            "Outlet Report",
            styles["Title"]
        )
    )

    elements.append(
        Paragraph(
            f"Generated: "
            f"{timezone.now().strftime('%B %d, %Y %I:%M %p')}",
            styles["Normal"]
        )
    )

    elements.append(Spacer(1, 20))

    # ========================================================
    # FILTER INFORMATION
    # ========================================================

    start = request.GET.get("start", "").strip()
    end = request.GET.get("end", "").strip()
    department = request.GET.get("department", "").strip()
    outlet = request.GET.get("outlet", "").strip()

    filter_text = []

    if outlet:
        outlet_obj = Outlet.objects.filter(
            id=outlet
        ).first()

        if outlet_obj:
            filter_text.append(
                f"Outlet: {outlet_obj.name}"
            )

    if department:
        filter_text.append(
            f"Department: {department}"
        )

    if start:
        filter_text.append(
            f"Start Date: {start}"
        )

    if end:
        filter_text.append(
            f"End Date: {end}"
        )

    if filter_text:

        elements.append(
            Paragraph(
                " | ".join(filter_text),
                styles["Normal"]
            )
        )

        elements.append(
            Spacer(1, 15)
        )

    # ========================================================
    # TABLE 1 — OUTLETS
    # ========================================================

    elements.append(
        Paragraph(
            "Outlets",
            styles["Heading2"]
        )
    )

    elements.append(
        Paragraph(
            "Ticket summary by outlet",
            styles["Normal"]
        )
    )

    elements.append(
        Spacer(1, 10)
    )

    outlet_data = [
        [
            "Outlet",
            "Total",
            "Pending",
            "Progress",
            "Resolved",
            "Cancelled",
        ]
    ]

    for row in outlet_summary:

        outlet_data.append([
            row["outlet__name"] or "N/A",
            row["total"],
            row["pending"],
            row["progress"],
            row["resolved"],
            row["cancelled"],
        ])

    outlet_table = Table(
        outlet_data,
        colWidths=[
            180,
            80,
            80,
            80,
            80,
            80,
        ],
        repeatRows=1,
    )

    outlet_table.setStyle(
        TableStyle([
            (
                "BACKGROUND",
                (0, 0),
                (-1, 0),
                colors.HexColor("#0f3554")
            ),

            (
                "TEXTCOLOR",
                (0, 0),
                (-1, 0),
                colors.white
            ),

            (
                "FONTNAME",
                (0, 0),
                (-1, 0),
                "Helvetica-Bold"
            ),

            (
                "FONTSIZE",
                (0, 0),
                (-1, -1),
                9
            ),

            (
                "GRID",
                (0, 0),
                (-1, -1),
                0.5,
                colors.HexColor("#dbe2ea")
            ),

            (
                "ROWBACKGROUNDS",
                (0, 1),
                (-1, -1),
                [
                    colors.white,
                    colors.HexColor("#f8fafc")
                ]
            ),

            (
                "VALIGN",
                (0, 0),
                (-1, -1),
                "MIDDLE"
            ),

            (
                "LEFTPADDING",
                (0, 0),
                (-1, -1),
                8
            ),

            (
                "RIGHTPADDING",
                (0, 0),
                (-1, -1),
                8
            ),

            (
                "TOPPADDING",
                (0, 0),
                (-1, -1),
                7
            ),

            (
                "BOTTOMPADDING",
                (0, 0),
                (-1, -1),
                7
            ),
        ])
    )

    elements.append(outlet_table)

    elements.append(
        Spacer(1, 30)
    )

    # ========================================================
    # TABLE 2 — CONCERNS PER OUTLET
    # ========================================================

    elements.append(
        Paragraph(
            "Concerns per Outlet",
            styles["Heading2"]
        )
    )

    elements.append(
        Paragraph(
            "Ticket concerns grouped by outlet",
            styles["Normal"]
        )
    )

    elements.append(
        Spacer(1, 10)
    )

    concern_data = [
        [
            "Outlet",
            "Concern",
            "Total",
        ]
    ]

    for row in concerns_per_outlet:

        concern_data.append([
            row["outlet__name"] or "N/A",
            row["concern_type__name"] or "N/A",
            row["total"],
        ])

    concern_table = Table(
        concern_data,
        colWidths=[
            220,
            300,
            100,
        ],
        repeatRows=1,
    )

    concern_table.setStyle(
        TableStyle([
            (
                "BACKGROUND",
                (0, 0),
                (-1, 0),
                colors.HexColor("#0f3554")
            ),

            (
                "TEXTCOLOR",
                (0, 0),
                (-1, 0),
                colors.white
            ),

            (
                "FONTNAME",
                (0, 0),
                (-1, 0),
                "Helvetica-Bold"
            ),

            (
                "FONTSIZE",
                (0, 0),
                (-1, -1),
                9
            ),

            (
                "GRID",
                (0, 0),
                (-1, -1),
                0.5,
                colors.HexColor("#dbe2ea")
            ),

            (
                "ROWBACKGROUNDS",
                (0, 1),
                (-1, -1),
                [
                    colors.white,
                    colors.HexColor("#f8fafc")
                ]
            ),

            (
                "VALIGN",
                (0, 0),
                (-1, -1),
                "MIDDLE"
            ),

            (
                "LEFTPADDING",
                (0, 0),
                (-1, -1),
                8
            ),

            (
                "RIGHTPADDING",
                (0, 0),
                (-1, -1),
                8
            ),

            (
                "TOPPADDING",
                (0, 0),
                (-1, -1),
                7
            ),

            (
                "BOTTOMPADDING",
                (0, 0),
                (-1, -1),
                7
            ),
        ])
    )

    elements.append(concern_table)

    # ========================================================
    # BUILD PDF
    # ========================================================

    doc.build(elements)

    return response


# ============================================================
# OUTLET EXPORT — EXCEL
# ============================================================

def export_outlets_excel(request):

    # Use EXACT SAME FILTERS as Outlet report page
    tickets = get_filtered_tickets(request)

    # ========================================================
    # OUTLET SUMMARY
    # ========================================================

    outlet_summary = (
        tickets
        .values("outlet__name")
        .annotate(
            total=Count("id"),

            pending=Count(
                "id",
                filter=Q(status="pending")
            ),

            progress=Count(
                "id",
                filter=Q(status="progress")
            ),

            resolved=Count(
                "id",
                filter=Q(status="resolved")
            ),

            cancelled=Count(
                "id",
                filter=Q(status="cancelled")
            ),
        )
        .order_by("outlet__name")
    )

    # ========================================================
    # CONCERNS PER OUTLET
    # ========================================================

    concerns_per_outlet = (
        tickets
        .values(
            "outlet__name",
            "concern_type__name"
        )
        .annotate(
            total=Count("id")
        )
        .order_by(
            "outlet__name",
            "-total"
        )
    )

    # ========================================================
    # WORKBOOK
    # ========================================================

    wb = Workbook()

    # ========================================================
    # SHEET 1 — OUTLETS
    # ========================================================

    ws1 = wb.active
    ws1.title = "Outlets"

    ws1.append([
        "Outlet",
        "Total",
        "Pending",
        "Progress",
        "Resolved",
        "Cancelled",
    ])

    for row in outlet_summary:

        ws1.append([
            row["outlet__name"] or "N/A",
            row["total"],
            row["pending"],
            row["progress"],
            row["resolved"],
            row["cancelled"],
        ])

    # ========================================================
    # SHEET 2 — CONCERNS PER OUTLET
    # ========================================================

    ws2 = wb.create_sheet(
        "Concerns per Outlet"
    )

    ws2.append([
        "Outlet",
        "Concern",
        "Total",
    ])

    for row in concerns_per_outlet:

        ws2.append([
            row["outlet__name"] or "N/A",
            row["concern_type__name"] or "N/A",
            row["total"],
        ])

    # ========================================================
    # EXCEL FORMATTING
    # ========================================================

    from openpyxl.styles import Font, PatternFill, Alignment

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="0F3554"
    )

    header_font = Font(
        bold=True,
        color="FFFFFF"
    )

    # OUTLETS SHEET

    for cell in ws1[1]:

        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center"
        )

    ws1.column_dimensions["A"].width = 25
    ws1.column_dimensions["B"].width = 12
    ws1.column_dimensions["C"].width = 12
    ws1.column_dimensions["D"].width = 12
    ws1.column_dimensions["E"].width = 12
    ws1.column_dimensions["F"].width = 12

    ws1.freeze_panes = "A2"

    # CONCERNS SHEET

    for cell in ws2[1]:

        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center"
        )

    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 35
    ws2.column_dimensions["C"].width = 12

    ws2.freeze_panes = "A2"

    # ========================================================
    # RESPONSE
    # ========================================================

    response = HttpResponse(
        content_type=(
            "application/"
            "vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    )

    selected_outlet = request.GET.get(
        "outlet",
        ""
    ).strip()

    selected_department = request.GET.get(
        "department",
        ""
    ).strip()

    if selected_outlet:

        outlet_obj = Outlet.objects.filter(
            id=selected_outlet
        ).first()

        outlet_name = (
            outlet_obj.name
            if outlet_obj
            else "Outlet"
        )

        filename = (
            f"{outlet_name}_outlet_report.xlsx"
            .replace(" ", "_")
        )

    elif selected_department:

        filename = (
            f"{selected_department}_outlet_report.xlsx"
            .replace(" ", "_")
        )

    else:

        filename = "outlets_report.xlsx"

    response["Content-Disposition"] = (
        f'attachment; filename="{filename}"'
    )

    wb.save(response)

    return response

# ============================================================
# DEPARTMENT REPORT DATA
# ============================================================

def get_department_report_data(request):

    # Use the EXACT SAME FILTERS
    # as the Department report page
    tickets = get_filtered_tickets(request).select_related(
        "department",
        "outlet",
        "concern_type",
    )

    departments = {}

    for ticket in tickets:

        if not ticket.department:
            continue

        department_name = ticket.department.name

        if department_name not in departments:

            departments[department_name] = {
                "total": 0,
                "open": 0,
                "progress": 0,
                "resolved": 0,
                "overdue": 0,
                "resolution_times": [],
            }

        data = departments[department_name]

        # TOTAL
        data["total"] += 1

        # STATUS
        if ticket.status == "pending":
            data["open"] += 1

        elif ticket.status == "progress":
            data["progress"] += 1

        elif ticket.status == "resolved":
            data["resolved"] += 1

        # OVERDUE
        if ticket.is_overdue:
            data["overdue"] += 1

        # RESOLUTION TIME
        if (
            ticket.status == "resolved"
            and ticket.resolve_at
            and ticket.created_at
        ):

            resolution_seconds = (
                ticket.resolve_at - ticket.created_at
            ).total_seconds()

            resolution_days = (
                resolution_seconds / 86400
            )

            data["resolution_times"].append(
                resolution_days
            )

    # ========================================================
    # FINALIZE PERFORMANCE DATA
    # ========================================================

    for department_name, data in departments.items():

        # Average resolution time
        if data["resolution_times"]:

            avg_resolution = (
                sum(data["resolution_times"])
                / len(data["resolution_times"])
            )

        else:

            avg_resolution = 0

        data["avg_resolution_time"] = avg_resolution

        # Resolution rate
        if data["total"]:

            data["resolution_rate"] = (
                data["resolved"]
                / data["total"]
            ) * 100

        else:

            data["resolution_rate"] = 0

    return departments

# ============================================================
# DEPARTMENT EXPORT — EXCEL
# ============================================================

def export_departments_excel(request):

    departments = get_department_report_data(request)

    wb = Workbook()

    # ========================================================
    # SHEET 1 — DEPARTMENTS
    # ========================================================

    ws1 = wb.active
    ws1.title = "Departments"

    ws1.append([
        "Department",
        "Total",
        "Open",
        "In Progress",
        "Resolved",
    ])

    for department_name in sorted(departments):

        data = departments[department_name]

        ws1.append([
            department_name,
            data["total"],
            data["open"],
            data["progress"],
            data["resolved"],
        ])

    # ========================================================
    # SHEET 2 — DEPARTMENT PERFORMANCE
    # ========================================================

    ws2 = wb.create_sheet(
        "Department Performance"
    )

    ws2.append([
        "Department",
        "Avg. Resolution Time",
        "Overdue",
        "Resolution Rate",
    ])

    for department_name in sorted(departments):

        data = departments[department_name]

        ws2.append([
            department_name,
            round(
                data["avg_resolution_time"],
                1
            ),
            data["overdue"],
            round(
                data["resolution_rate"],
                1
            ) / 100,
        ])

    # ========================================================
    # EXCEL STYLING
    # ========================================================

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="4F46E5"
    )

    header_font = Font(
        bold=True,
        color="FFFFFF"
    )

    # --------------------------------------------------------
    # SHEET 1
    # --------------------------------------------------------

    for cell in ws1[1]:

        cell.fill = header_fill
        cell.font = header_font

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

    ws1.column_dimensions["A"].width = 25
    ws1.column_dimensions["B"].width = 12
    ws1.column_dimensions["C"].width = 12
    ws1.column_dimensions["D"].width = 15
    ws1.column_dimensions["E"].width = 12

    ws1.freeze_panes = "A2"

    # --------------------------------------------------------
    # SHEET 2
    # --------------------------------------------------------

    for cell in ws2[1]:

        cell.fill = header_fill
        cell.font = header_font

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 25
    ws2.column_dimensions["C"].width = 12
    ws2.column_dimensions["D"].width = 18

    ws2.freeze_panes = "A2"

    # Resolution rate as percentage
    for row in range(2, ws2.max_row + 1):

        ws2.cell(
            row=row,
            column=4
        ).number_format = "0.0%"

    # ========================================================
    # FILTER INFORMATION
    # ========================================================

    selected_department = request.GET.get(
        "department",
        ""
    ).strip()

    selected_outlet = request.GET.get(
        "outlet",
        ""
    ).strip()

    # ========================================================
    # FILE NAME
    # ========================================================

    if selected_department:

        filename = (
            f"{selected_department}"
            "_department_report.xlsx"
        )

    elif selected_outlet:

        outlet_obj = Outlet.objects.filter(
            id=selected_outlet
        ).first()

        outlet_name = (
            outlet_obj.name
            if outlet_obj
            else "Outlet"
        )

        filename = (
            f"{outlet_name}"
            "_department_report.xlsx"
        )

    else:

        filename = "departments_report.xlsx"

    filename = filename.replace(
        " ",
        "_"
    )

    # ========================================================
    # RESPONSE
    # ========================================================

    response = HttpResponse(
        content_type=(
            "application/"
            "vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    )

    response["Content-Disposition"] = (
        f'attachment; filename="{filename}"'
    )

    wb.save(response)

    return response

# ============================================================
# DEPARTMENT EXPORT — PDF
# ============================================================

def export_departments_pdf(request):

    departments = get_department_report_data(request)

    response = HttpResponse(
        content_type="application/pdf"
    )

    # ========================================================
    # FILE NAME
    # ========================================================

    selected_department = request.GET.get(
        "department",
        ""
    ).strip()

    selected_outlet = request.GET.get(
        "outlet",
        ""
    ).strip()

    if selected_department:

        filename = (
            f"{selected_department}"
            "_department_report.pdf"
        )

    elif selected_outlet:

        outlet_obj = Outlet.objects.filter(
            id=selected_outlet
        ).first()

        outlet_name = (
            outlet_obj.name
            if outlet_obj
            else "Outlet"
        )

        filename = (
            f"{outlet_name}"
            "_department_report.pdf"
        )

    else:

        filename = "departments_report.pdf"

    filename = filename.replace(
        " ",
        "_"
    )

    response["Content-Disposition"] = (
        f'attachment; filename="{filename}"'
    )

    # ========================================================
    # DOCUMENT
    # ========================================================

    document = SimpleDocTemplate(
        response,
        pagesize=landscape(A4),

        rightMargin=30,
        leftMargin=30,

        topMargin=30,
        bottomMargin=30,
    )

    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        "DepartmentReportTitle",

        parent=styles["Title"],

        fontSize=18,

        alignment=TA_CENTER,

        spaceAfter=8,
    )

    subtitle_style = ParagraphStyle(
        "DepartmentReportSubtitle",

        parent=styles["Normal"],

        fontSize=9,

        textColor=colors.HexColor(
            "#64748B"
        ),

        spaceAfter=15,
    )

    elements = []

    # ========================================================
    # TITLE
    # ========================================================

    elements.append(
        Paragraph(
            "FB Management — Department Report",
            title_style
        )
    )

    elements.append(
        Paragraph(
            "Generated: "
            + timezone.now().strftime(
                "%B %d, %Y %I:%M %p"
            ),
            subtitle_style
        )
    )

    # ========================================================
    # FILTER INFORMATION
    # ========================================================

    start = request.GET.get(
        "start",
        ""
    ).strip()

    end = request.GET.get(
        "end",
        ""
    ).strip()

    department = request.GET.get(
        "department",
        ""
    ).strip()

    outlet = request.GET.get(
        "outlet",
        ""
    ).strip()

    filter_text = []

    if department:

        filter_text.append(
            f"Department: {department}"
        )

    if outlet:

        outlet_obj = Outlet.objects.filter(
            id=outlet
        ).first()

        if outlet_obj:

            filter_text.append(
                f"Outlet: {outlet_obj.name}"
            )

    if start:

        filter_text.append(
            f"Start Date: {start}"
        )

    if end:

        filter_text.append(
            f"End Date: {end}"
        )

    if filter_text:

        elements.append(
            Paragraph(
                " | ".join(filter_text),
                styles["Normal"]
            )
        )

        elements.append(
            Spacer(1, 15)
        )

    # ========================================================
    # TABLE 1 — DEPARTMENTS
    # ========================================================

    elements.append(
        Paragraph(
            "Departments",
            styles["Heading2"]
        )
    )

    elements.append(
        Paragraph(
            "Ticket summary by department",
            styles["Normal"]
        )
    )

    elements.append(
        Spacer(1, 10)
    )

    department_data = [
        [
            "Department",
            "Total",
            "Open",
            "In Progress",
            "Resolved",
        ]
    ]

    for department_name in sorted(
        departments
    ):

        data = departments[
            department_name
        ]

        department_data.append([
            department_name,
            data["total"],
            data["open"],
            data["progress"],
            data["resolved"],
        ])

    department_table = Table(
        department_data,

        colWidths=[
            230,
            80,
            80,
            100,
            80,
        ],

        repeatRows=1,
    )

    department_table.setStyle(
        TableStyle([

            (
                "BACKGROUND",
                (0, 0),
                (-1, 0),
                colors.HexColor("#4F46E5")
            ),

            (
                "TEXTCOLOR",
                (0, 0),
                (-1, 0),
                colors.white
            ),

            (
                "FONTNAME",
                (0, 0),
                (-1, 0),
                "Helvetica-Bold"
            ),

            (
                "FONTSIZE",
                (0, 0),
                (-1, -1),
                9
            ),

            (
                "GRID",
                (0, 0),
                (-1, -1),
                0.5,
                colors.HexColor("#E2E8F0")
            ),

            (
                "ROWBACKGROUNDS",
                (0, 1),
                (-1, -1),
                [
                    colors.white,
                    colors.HexColor("#F8FAFC")
                ]
            ),

            (
                "VALIGN",
                (0, 0),
                (-1, -1),
                "MIDDLE"
            ),

            (
                "LEFTPADDING",
                (0, 0),
                (-1, -1),
                8
            ),

            (
                "RIGHTPADDING",
                (0, 0),
                (-1, -1),
                8
            ),

            (
                "TOPPADDING",
                (0, 0),
                (-1, -1),
                7
            ),

            (
                "BOTTOMPADDING",
                (0, 0),
                (-1, -1),
                7
            ),
        ])
    )

    elements.append(
        department_table
    )

    elements.append(
        Spacer(1, 30)
    )

    # ========================================================
    # TABLE 2 — DEPARTMENT PERFORMANCE
    # ========================================================

    elements.append(
        Paragraph(
            "Department Performance",
            styles["Heading2"]
        )
    )

    elements.append(
        Paragraph(
            "Resolution performance by department",
            styles["Normal"]
        )
    )

    elements.append(
        Spacer(1, 10)
    )

    performance_data = [
        [
            "Department",
            "Avg. Resolution Time",
            "Overdue",
            "Resolution Rate",
        ]
    ]

    for department_name in sorted(
        departments
    ):

        data = departments[
            department_name
        ]

        avg_days = data[
            "avg_resolution_time"
        ]

        resolution_rate = data[
            "resolution_rate"
        ]

        performance_data.append([
            department_name,

            f"{avg_days:.1f} days",

            data["overdue"],

            f"{resolution_rate:.1f}%",
        ])

    performance_table = Table(
        performance_data,

        colWidths=[
            230,
            180,
            100,
            150,
        ],

        repeatRows=1,
    )

    performance_table.setStyle(
        TableStyle([

            (
                "BACKGROUND",
                (0, 0),
                (-1, 0),
                colors.HexColor("#4F46E5")
            ),

            (
                "TEXTCOLOR",
                (0, 0),
                (-1, 0),
                colors.white
            ),

            (
                "FONTNAME",
                (0, 0),
                (-1, 0),
                "Helvetica-Bold"
            ),

            (
                "FONTSIZE",
                (0, 0),
                (-1, -1),
                9
            ),

            (
                "GRID",
                (0, 0),
                (-1, -1),
                0.5,
                colors.HexColor("#E2E8F0")
            ),

            (
                "ROWBACKGROUNDS",
                (0, 1),
                (-1, -1),
                [
                    colors.white,
                    colors.HexColor("#F8FAFC")
                ]
            ),

            (
                "VALIGN",
                (0, 0),
                (-1, -1),
                "MIDDLE"
            ),

            (
                "LEFTPADDING",
                (0, 0),
                (-1, -1),
                8
            ),

            (
                "RIGHTPADDING",
                (0, 0),
                (-1, -1),
                8
            ),

            (
                "TOPPADDING",
                (0, 0),
                (-1, -1),
                7
            ),

            (
                "BOTTOMPADDING",
                (0, 0),
                (-1, -1),
                7
            ),
        ])
    )

    elements.append(
        performance_table
    )

    # ========================================================
    # BUILD
    # ========================================================

    document.build(
        elements
    )

    return response

# ============================================================
# TECHNICIAN REPORT DATA
# ============================================================

def get_technician_report_data(request):

    tickets = get_filtered_tickets(request)

    technicians = Technician.objects.all().order_by("full_name")

    # --------------------------------------------------------
    # DEPARTMENT FILTER
    # --------------------------------------------------------

    department = request.GET.get("department", "").strip()

    if department:
        technicians = technicians.filter(
            department__name=department
        )

    now = timezone.now()

    technician_rows = []
    technician_performance = []

    # ========================================================
    # EACH TECHNICIAN
    # ========================================================

    for tech in technicians:

        # ----------------------------------------------------
        # TICKETS CURRENTLY ASSIGNED
        #
        # Includes:
        # 1. Primary technician
        # 2. Additional technicians
        # ----------------------------------------------------

        assigned_tickets = tickets.filter(
            Q(assigned_to=tech) |
            Q(additional_technicians=tech)
        ).distinct()

        # ----------------------------------------------------
        # TOTAL CURRENT WORKLOAD
        # ----------------------------------------------------

        total = assigned_tickets.count()

        # ----------------------------------------------------
        # STATUS COUNTS
        # ----------------------------------------------------

        pending = assigned_tickets.filter(
            status="pending"
        ).count()

        progress = assigned_tickets.filter(
            status="progress"
        ).count()

        resolved = assigned_tickets.filter(
            status="resolved"
        ).count()

        cancelled = assigned_tickets.filter(
            status="cancelled"
        ).count()

        # ----------------------------------------------------
        # OVERDUE
        # ----------------------------------------------------

        overdue = assigned_tickets.filter(
            deadline__lt=now
        ).exclude(
            status__in=[
                "resolved",
                "cancelled",
            ]
        ).count()

        # ----------------------------------------------------
        # TOTAL ASSIGNMENTS
        #
        # Uses your existing assignment logs.
        # ----------------------------------------------------

        primary_assignments = (
            TicketAssignmentLog.objects
            .filter(
                new_technician=tech,
                ticket__in=tickets,
            )
            .count()
        )

        additional_assignments = (
            TicketAdditionalAssignmentLog.objects
            .filter(
                technician=tech,
                action="added",
                ticket__in=tickets,
            )
            .count()
        )

        total_assigned = (
            primary_assignments +
            additional_assignments
        )

        # ----------------------------------------------------
        # RESOLUTION TIME
        # ----------------------------------------------------

        resolved_tickets = assigned_tickets.filter(
            status="resolved",
            resolve_at__isnull=False,
            created_at__isnull=False,
        )

        resolution_seconds = []

        for ticket in resolved_tickets:

            if (
                ticket.resolve_at and
                ticket.created_at
            ):

                duration = (
                    ticket.resolve_at -
                    ticket.created_at
                )

                resolution_seconds.append(
                    duration.total_seconds()
                )

        if resolution_seconds:

            average_seconds = (
                sum(resolution_seconds) /
                len(resolution_seconds)
            )

            average_days = round(
                average_seconds / 86400,
                1
            )

        else:

            average_days = 0

        # ----------------------------------------------------
        # RESOLUTION RATE
        # ----------------------------------------------------

        if total > 0:

            resolution_rate = (
                resolved / total
            ) * 100

        else:

            resolution_rate = 0

        resolution_rate = round(
            resolution_rate,
            1
        )

        # ----------------------------------------------------
        # ON-TIME RESOLUTIONS
        # ----------------------------------------------------

        on_time = assigned_tickets.filter(
            status="resolved",
            resolve_at__isnull=False,
            deadline__isnull=False,
            resolve_at__lte=F("deadline"),
        ).count()

        # ----------------------------------------------------
        # REOPENED
        # ----------------------------------------------------

        reopened = TicketStatusLog.objects.filter(
            technician=tech,
            old_status="resolved",
            ticket__in=tickets,
        ).count()

        # ====================================================
        # TABLE 1
        # ====================================================

        technician_rows.append({

            "name":
                tech.full_name or "Unnamed Technician",

            "total":
                total,

            "open":
                pending,

            "progress":
                progress,

            "resolved":
                resolved,

            "cancelled":
                cancelled,

            "total_assigned":
                total_assigned,

        })

        # ====================================================
        # TABLE 2
        # ====================================================

        technician_performance.append({

            "name":
                tech.full_name or "Unnamed Technician",

            "average_days":
                average_days,

            "overdue":
                overdue,

            "resolution_rate":
                resolution_rate,

            "on_time":
                on_time,

            "reopened":
                reopened,

            "total":
                total,

            "resolved":
                resolved,

        })

    return (
        technician_rows,
        technician_performance,
    )
# ============================================================
# TECHNICIAN REPORT — PDF
# ============================================================

def export_technicians_pdf(request):

    (
        technician_rows,
        technician_performance,
    ) = get_technician_report_data(request)

    response = HttpResponse(
        content_type="application/pdf"
    )

    # --------------------------------------------------------
    # FILENAME
    # --------------------------------------------------------

    selected_department = (
        request.GET.get(
            "department",
            ""
        ).strip()
    )

    selected_outlet = (
        request.GET.get(
            "outlet",
            ""
        ).strip()
    )

    if selected_department:

        filename = (
            f"{selected_department}_technician_report.pdf"
            .replace(" ", "_")
        )

    elif selected_outlet:

        outlet_obj = Outlet.objects.filter(
            id=selected_outlet
        ).first()

        outlet_name = (
            outlet_obj.name
            if outlet_obj
            else "Outlet"
        )

        filename = (
            f"{outlet_name}_technician_report.pdf"
            .replace(" ", "_")
        )

    else:

        filename = "technicians_report.pdf"

    response["Content-Disposition"] = (
        f'attachment; filename="{filename}"'
    )

    # ========================================================
    # PDF
    # ========================================================

    doc = SimpleDocTemplate(
        response,
        pagesize=landscape(A4),
        rightMargin=30,
        leftMargin=30,
        topMargin=30,
        bottomMargin=30,
    )

    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        "TechnicianReportTitle",
        parent=styles["Title"],
        fontSize=18,
        alignment=TA_CENTER,
        spaceAfter=15,
    )

    elements = []

    # ========================================================
    # TITLE
    # ========================================================

    elements.append(
        Paragraph(
            "FB Management — Technician Report",
            title_style
        )
    )

    elements.append(
        Paragraph(
            f"Generated: "
            f"{timezone.now().strftime('%B %d, %Y %I:%M %p')}",
            styles["Normal"]
        )
    )

    elements.append(
        Spacer(1, 12)
    )

    # ========================================================
    # FILTER INFORMATION
    # ========================================================

    start = request.GET.get(
        "start",
        ""
    ).strip()

    end = request.GET.get(
        "end",
        ""
    ).strip()

    department = request.GET.get(
        "department",
        ""
    ).strip()

    outlet = request.GET.get(
        "outlet",
        ""
    ).strip()

    filter_parts = []

    if outlet:

        outlet_obj = Outlet.objects.filter(
            id=outlet
        ).first()

        if outlet_obj:

            filter_parts.append(
                f"Outlet: {outlet_obj.name}"
            )

    else:

        filter_parts.append(
            "Outlet: All Outlets"
        )

    filter_parts.append(
        f"Department: "
        f"{department or 'All Departments'}"
    )

    filter_parts.append(
        f"Date: "
        f"{start or 'All'}"
        f" to "
        f"{end or 'All'}"
    )

    elements.append(
        Paragraph(
            " | ".join(filter_parts),
            styles["Normal"]
        )
    )

    elements.append(
        Spacer(1, 20)
    )

    # ========================================================
    # TABLE 1 — TECHNICIAN WORKLOAD
    # ========================================================

    elements.append(
        Paragraph(
            "Technician Workload",
            styles["Heading2"]
        )
    )

    elements.append(
        Paragraph(
            "Ticket volume and current workload by technician.",
            styles["Normal"]
        )
    )

    elements.append(
        Spacer(1, 10)
    )

    workload_data = [[
        "Technician",
        "Total",
        "Pending",
        "In Progress",
        "Resolved",
        "Cancelled",
    ]]

    for tech in technician_rows:

        workload_data.append([
            tech["name"],
            tech["total"],
            tech["open"],
            tech["progress"],
            tech["resolved"],
            tech["cancelled"],
        ])

    workload_table = Table(
        workload_data,
        repeatRows=1,
        colWidths=[
            190,
            70,
            80,
            100,
            80,
            80,
        ],
    )

    workload_table.setStyle(
        TableStyle([

            (
                "BACKGROUND",
                (0, 0),
                (-1, 0),
                colors.HexColor("#4f46e5")
            ),

            (
                "TEXTCOLOR",
                (0, 0),
                (-1, 0),
                colors.white
            ),

            (
                "FONTNAME",
                (0, 0),
                (-1, 0),
                "Helvetica-Bold"
            ),

            (
                "FONTSIZE",
                (0, 0),
                (-1, -1),
                9
            ),

            (
                "GRID",
                (0, 0),
                (-1, -1),
                0.5,
                colors.HexColor("#e5e7eb")
            ),

            (
                "ROWBACKGROUNDS",
                (0, 1),
                (-1, -1),
                [
                    colors.white,
                    colors.HexColor("#f8fafc")
                ]
            ),

            (
                "VALIGN",
                (0, 0),
                (-1, -1),
                "MIDDLE"
            ),

            (
                "TOPPADDING",
                (0, 0),
                (-1, -1),
                8
            ),

            (
                "BOTTOMPADDING",
                (0, 0),
                (-1, -1),
                8
            ),

        ])
    )

    elements.append(
        workload_table
    )

    elements.append(
        Spacer(1, 25)
    )

    # ========================================================
    # TABLE 2 — TECHNICIAN PERFORMANCE
    # ========================================================

    elements.append(
        Paragraph(
            "Technician Performance",
            styles["Heading2"]
        )
    )

    elements.append(
        Paragraph(
            "Resolution performance and efficiency by technician.",
            styles["Normal"]
        )
    )

    elements.append(
        Spacer(1, 10)
    )

    performance_data = [[
        "Technician",
        "Avg. Resolution Time",
        "Overdue",
        "Resolution Rate",
        "On Time",
        "Reopened",
    ]]

    for tech in technician_performance:

        performance_data.append([
            tech["name"],
            f'{tech["average_days"]} days',
            tech["overdue"],
            f'{tech["resolution_rate"]:.1f}%',
            tech["on_time"],
            tech["reopened"],
        ])

    performance_table = Table(
        performance_data,
        repeatRows=1,
        colWidths=[
            170,
            125,
            75,
            100,
            70,
            75,
        ],
    )

    performance_table.setStyle(
        TableStyle([

            (
                "BACKGROUND",
                (0, 0),
                (-1, 0),
                colors.HexColor("#4f46e5")
            ),

            (
                "TEXTCOLOR",
                (0, 0),
                (-1, 0),
                colors.white
            ),

            (
                "FONTNAME",
                (0, 0),
                (-1, 0),
                "Helvetica-Bold"
            ),

            (
                "FONTSIZE",
                (0, 0),
                (-1, -1),
                9
            ),

            (
                "GRID",
                (0, 0),
                (-1, -1),
                0.5,
                colors.HexColor("#e5e7eb")
            ),

            (
                "ROWBACKGROUNDS",
                (0, 1),
                (-1, -1),
                [
                    colors.white,
                    colors.HexColor("#f8fafc")
                ]
            ),

            (
                "VALIGN",
                (0, 0),
                (-1, -1),
                "MIDDLE"
            ),

            (
                "TOPPADDING",
                (0, 0),
                (-1, -1),
                8
            ),

            (
                "BOTTOMPADDING",
                (0, 0),
                (-1, -1),
                8
            ),

        ])
    )

    elements.append(
        performance_table
    )

    doc.build(elements)

    return response

# ============================================================
# TECHNICIAN REPORT — EXCEL
# ============================================================

def export_technicians_excel(request):

    (
        technician_rows,
        technician_performance,
    ) = get_technician_report_data(request)

    # ========================================================
    # WORKBOOK
    # ========================================================

    wb = Workbook()

    # ========================================================
    # SHEET 1 — TECHNICIAN WORKLOAD
    # ========================================================

    ws1 = wb.active

    ws1.title = "Technician Workload"

    ws1.append([
        "Technician",
        "Total",
        "Pending",
        "In Progress",
        "Resolved",
        "Cancelled",
        "Total Assignments",
    ])

    for tech in technician_rows:

        ws1.append([
            tech["name"],
            tech["total"],
            tech["open"],
            tech["progress"],
            tech["resolved"],
            tech["cancelled"],
            tech["total_assigned"],
        ])

    # ========================================================
    # SHEET 2 — TECHNICIAN PERFORMANCE
    # ========================================================

    ws2 = wb.create_sheet(
        "Technician Performance"
    )

    ws2.append([
        "Technician",
        "Avg. Resolution Time (Days)",
        "Overdue",
        "Resolution Rate",
        "On Time",
        "Reopened",
        "Total",
        "Resolved",
    ])

    for tech in technician_performance:

        ws2.append([
            tech["name"],
            tech["average_days"],
            tech["overdue"],
            tech["resolution_rate"],
            tech["on_time"],
            tech["reopened"],
            tech["total"],
            tech["resolved"],
        ])

    # ========================================================
    # EXCEL STYLE
    # ========================================================

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="4F46E5"
    )

    header_font = Font(
        bold=True,
        color="FFFFFF"
    )

    thin_border = Border(
        bottom=Side(
            style="thin",
            color="E5E7EB"
        )
    )

    # --------------------------------------------------------
    # WORKLOAD SHEET
    # --------------------------------------------------------

    for cell in ws1[1]:

        cell.fill = header_fill

        cell.font = header_font

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

        cell.border = thin_border

    ws1.freeze_panes = "A2"

    ws1.auto_filter.ref = (
        ws1.dimensions
    )

    # --------------------------------------------------------
    # PERFORMANCE SHEET
    # --------------------------------------------------------

    for cell in ws2[1]:

        cell.fill = header_fill

        cell.font = header_font

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

        cell.border = thin_border

    ws2.freeze_panes = "A2"

    ws2.auto_filter.ref = (
        ws2.dimensions
    )

    # ========================================================
    # AUTO WIDTH
    # ========================================================

    for ws in [ws1, ws2]:

        for column_cells in ws.columns:

            max_length = 0

            column_letter = (
                get_column_letter(
                    column_cells[0].column
                )
            )

            for cell in column_cells:

                if cell.value is not None:

                    max_length = max(
                        max_length,
                        len(str(cell.value))
                    )

            ws.column_dimensions[
                column_letter
            ].width = min(
                max_length + 3,
                35
            )

    # ========================================================
    # FILTER INFORMATION
    # ========================================================

    ws1.insert_rows(1, 2)
    ws2.insert_rows(1, 2)

    ws1["A1"] = (
        "FB Management — Technician Report"
    )

    ws2["A1"] = (
        "FB Management — Technician Performance"
    )

    ws1["A1"].font = Font(
        bold=True,
        size=16
    )

    ws2["A1"].font = Font(
        bold=True,
        size=16
    )

    ws1.merge_cells(
        start_row=1,
        start_column=1,
        end_row=1,
        end_column=7
    )

    ws2.merge_cells(
        start_row=1,
        start_column=1,
        end_row=1,
        end_column=8
    )

    # ========================================================
    # RESPONSE
    # ========================================================

    response = HttpResponse(
        content_type=(
            "application/"
            "vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    )

    selected_department = (
        request.GET.get(
            "department",
            ""
        ).strip()
    )

    selected_outlet = (
        request.GET.get(
            "outlet",
            ""
        ).strip()
    )

    if selected_department:

        filename = (
            f"{selected_department}_"
            f"technician_report.xlsx"
        )

    elif selected_outlet:

        outlet_obj = Outlet.objects.filter(
            id=selected_outlet
        ).first()

        outlet_name = (
            outlet_obj.name
            if outlet_obj
            else "Outlet"
        )

        filename = (
            f"{outlet_name}_"
            f"technician_report.xlsx"
            .replace(" ", "_")
        )

    else:

        filename = "technicians_report.xlsx"

    response["Content-Disposition"] = (
        f'attachment; filename="{filename}"'
    )

    wb.save(response)

    return response

# =========================================================
# CONCERN REPORT DATA
# =========================================================

def get_concern_report_data(request):

    selected_start = request.GET.get(
        "start", ""
    ).strip()

    selected_end = request.GET.get(
        "end", ""
    ).strip()

    selected_department = request.GET.get(
        "department", ""
    ).strip()

    selected_outlet = request.GET.get(
        "outlet", ""
    ).strip()

    selected_concern = request.GET.get(
        "concern", ""
    ).strip()


    # =====================================================
    # BASE QUERY
    # =====================================================

    tickets = Ticket.objects.all()


    # =====================================================
    # START DATE
    # =====================================================

    if selected_start:

        try:

            start_date = datetime.strptime(
                selected_start,
                "%Y-%m-%d"
            ).date()

            tickets = tickets.filter(
                created_at__date__gte=start_date
            )

        except ValueError:

            pass


    # =====================================================
    # END DATE
    # =====================================================

    if selected_end:

        try:

            end_date = datetime.strptime(
                selected_end,
                "%Y-%m-%d"
            ).date()

            tickets = tickets.filter(
                created_at__date__lte=end_date
            )

        except ValueError:

            pass


    # =====================================================
    # DEPARTMENT
    # =====================================================

    if selected_department:

        tickets = tickets.filter(
            department__name=selected_department
        )


    # =====================================================
    # OUTLET
    # =====================================================

    if selected_outlet:

        tickets = tickets.filter(
            outlet_id=selected_outlet
        )


    # =====================================================
    # CONCERN
    # =====================================================

    if selected_concern:

        tickets = tickets.filter(
            concern_type_id=selected_concern
        )


    # =====================================================
    # SUMMARY
    # =====================================================

    concern_summary = (
        tickets
        .values(
            "concern_type_id",
            "concern_type__name",
            "department__name",
        )
        .annotate(

            total=Count("id"),

            open=Count(
                "id",
                filter=Q(status="pending")
            ),

            progress=Count(
                "id",
                filter=Q(status="progress")
            ),

            resolved=Count(
                "id",
                filter=Q(status="resolved")
            ),

        )
        .order_by(
            "concern_type__name"
        )
    )


    # =====================================================
    # PERFORMANCE
    # =====================================================

    performance = []


    for row in concern_summary:

        concern_id = row["concern_type_id"]


        if not concern_id:
            continue


        concern_tickets = tickets.filter(
            concern_type_id=concern_id
        )


        resolved_tickets = concern_tickets.filter(
            status="resolved",
            resolve_at__isnull=False,
            created_at__isnull=False,
        )


        # =================================================
        # AVG RESOLUTION
        # =================================================

        resolution_seconds = []


        for ticket in resolved_tickets:

            duration = (
                ticket.resolve_at -
                ticket.created_at
            )

            resolution_seconds.append(
                duration.total_seconds()
            )


        if resolution_seconds:

            average_seconds = (
                sum(resolution_seconds)
                /
                len(resolution_seconds)
            )

            average_days = round(
                average_seconds / 86400,
                1
            )

        else:

            average_days = 0


        # =================================================
        # OVERDUE
        #
        # FIX FOR YOUR CURRENT ERROR:
        # F("deadline")
        # =================================================

        overdue = concern_tickets.filter(
            deadline__isnull=False,
            resolve_at__isnull=False,
            resolve_at__lte=F("deadline"),
            status="resolved",
        ).count()


        # =================================================
        # RESOLUTION RATE
        # =================================================

        total_count = concern_tickets.count()

        resolved_count = resolved_tickets.count()


        if total_count:

            resolution_rate = round(
                resolved_count /
                total_count *
                100,
                1
            )

        else:

            resolution_rate = 0


        performance.append({

            "name":
                row["concern_type__name"],

            "department":
                row["department__name"] or "—",

            "total":
                row["total"],

            "open":
                row["open"],

            "progress":
                row["progress"],

            "resolved":
                row["resolved"],

            "average_days":
                average_days,

            "overdue":
                overdue,

            "resolution_rate":
                resolution_rate,

        })


    return (
        performance,
        selected_start,
        selected_end,
        selected_department,
        selected_outlet,
        selected_concern,
    )

def export_concerns_pdf(request):

    (
        concerns,
        selected_start,
        selected_end,
        selected_department,
        selected_outlet,
        selected_concern,
    ) = get_concern_report_data(request)


    response = HttpResponse(
        content_type="application/pdf"
    )

    response[
        "Content-Disposition"
    ] = 'attachment; filename="concern_report.pdf"'


    document = SimpleDocTemplate(
        response,
        pagesize=landscape(A4),

        rightMargin=30,
        leftMargin=30,
        topMargin=30,
        bottomMargin=30,
    )


    styles = getSampleStyleSheet()


    title_style = styles["Title"]

    title_style.alignment = TA_CENTER


    normal_style = styles["Normal"]


    story = []


    # =====================================================
    # TITLE
    # =====================================================

    story.append(
        Paragraph(
            "FB MANAGEMENT GROUP INC.",
            title_style
        )
    )


    story.append(
        Paragraph(
            "Concern Report",
            styles["Heading2"]
        )
    )


    story.append(
        Spacer(1, 10)
    )


    # =====================================================
    # FILTER INFORMATION
    # =====================================================

    filter_text = (
        f"Start: {selected_start or 'All'} "
        f"&nbsp;&nbsp;&nbsp; "
        f"End: {selected_end or 'All'} "
        f"&nbsp;&nbsp;&nbsp; "
        f"Department: {selected_department or 'All'} "
        f"&nbsp;&nbsp;&nbsp; "
        f"Outlet: {selected_outlet or 'All'}"
    )


    story.append(
        Paragraph(
            filter_text,
            normal_style
        )
    )


    story.append(
        Spacer(1, 15)
    )


    # =====================================================
    # SUMMARY TABLE
    # =====================================================

    summary_data = [

        [
            "Concern",
            "Department",
            "Total",
            "Open",
            "In Progress",
            "Resolved",
        ]

    ]


    for row in concerns:

        summary_data.append([

            row["name"] or "—",

            row["department"] or "—",

            str(row["total"]),

            str(row["open"]),

            str(row["progress"]),

            str(row["resolved"]),

        ])


    if len(summary_data) == 1:

        summary_data.append([
            "No data",
            "",
            "",
            "",
            "",
            "",
        ])


    table = Table(
        summary_data,
        repeatRows=1,
        colWidths=[
            150,
            110,
            70,
            70,
            90,
            70,
        ],
    )


    table.setStyle(
        TableStyle([

            (
                "BACKGROUND",
                (0, 0),
                (-1, 0),
                colors.HexColor("#4F46E5")
            ),

            (
                "TEXTCOLOR",
                (0, 0),
                (-1, 0),
                colors.white
            ),

            (
                "FONTNAME",
                (0, 0),
                (-1, 0),
                "Helvetica-Bold"
            ),

            (
                "FONTNAME",
                (0, 1),
                (-1, -1),
                "Helvetica"
            ),

            (
                "FONTSIZE",
                (0, 0),
                (-1, -1),
                8
            ),

            (
                "GRID",
                (0, 0),
                (-1, -1),
                0.5,
                colors.HexColor("#D9DCE3")
            ),

            (
                "VALIGN",
                (0, 0),
                (-1, -1),
                "MIDDLE"
            ),

            (
                "ALIGN",
                (2, 1),
                (-1, -1),
                "CENTER"
            ),

            (
                "TOPPADDING",
                (0, 0),
                (-1, -1),
                7
            ),

            (
                "BOTTOMPADDING",
                (0, 0),
                (-1, -1),
                7
            ),

        ])
    )


    story.append(
        Paragraph(
            "Concern Summary",
            styles["Heading2"]
        )
    )

    story.append(table)


    story.append(
        Spacer(1, 20)
    )


    # =====================================================
    # PERFORMANCE TABLE
    # =====================================================

    performance_data = [

        [
            "Concern",
            "Avg. Resolution Time",
            "Overdue",
            "Resolution Rate",
        ]

    ]


    for row in concerns:

        performance_data.append([

            row["name"] or "—",

            f'{row["average_days"]} day'
            + (
                "s"
                if row["average_days"] != 1
                else ""
            ),

            str(row["overdue"]),

            f'{row["resolution_rate"]:.1f}%',

        ])


    if len(performance_data) == 1:

        performance_data.append([
            "No data",
            "",
            "",
            "",
        ])


    performance_table = Table(
        performance_data,
        repeatRows=1,
        colWidths=[
            220,
            180,
            100,
            150,
        ],
    )


    performance_table.setStyle(
        TableStyle([

            (
                "BACKGROUND",
                (0, 0),
                (-1, 0),
                colors.HexColor("#4F46E5")
            ),

            (
                "TEXTCOLOR",
                (0, 0),
                (-1, 0),
                colors.white
            ),

            (
                "FONTNAME",
                (0, 0),
                (-1, 0),
                "Helvetica-Bold"
            ),

            (
                "FONTSIZE",
                (0, 0),
                (-1, -1),
                8
            ),

            (
                "GRID",
                (0, 0),
                (-1, -1),
                0.5,
                colors.HexColor("#D9DCE3")
            ),

            (
                "ALIGN",
                (2, 1),
                (-1, -1),
                "CENTER"
            ),

            (
                "TOPPADDING",
                (0, 0),
                (-1, -1),
                7
            ),

            (
                "BOTTOMPADDING",
                (0, 0),
                (-1, -1),
                7
            ),

        ])
    )


    story.append(
        Paragraph(
            "Concern Performance",
            styles["Heading2"]
        )
    )

    story.append(
        performance_table
    )


    document.build(story)


    return response

def export_concerns_excel(request):

    (
        concerns,
        selected_start,
        selected_end,
        selected_department,
        selected_outlet,
        selected_concern,
    ) = get_concern_report_data(request)


    workbook = Workbook()


    # =====================================================
    # SHEET 1 — SUMMARY
    # =====================================================

    worksheet = workbook.active

    worksheet.title = "Concern Summary"


    headers = [

        "Concern",
        "Department",
        "Total",
        "Open",
        "In Progress",
        "Resolved",

    ]


    worksheet.append(headers)


    for cell in worksheet[1]:

        cell.font = Font(
            bold=True,
            color="FFFFFF"
        )

        cell.fill = PatternFill(
            fill_type="solid",
            fgColor="4F46E5"
        )

        cell.alignment = Alignment(
            horizontal="center"
        )


    for row in concerns:

        worksheet.append([

            row["name"] or "—",

            row["department"] or "—",

            row["total"],

            row["open"],

            row["progress"],

            row["resolved"],

        ])


    # =====================================================
    # COLUMN WIDTH
    # =====================================================

    worksheet.column_dimensions["A"].width = 35
    worksheet.column_dimensions["B"].width = 25
    worksheet.column_dimensions["C"].width = 12
    worksheet.column_dimensions["D"].width = 12
    worksheet.column_dimensions["E"].width = 15
    worksheet.column_dimensions["F"].width = 12


    # =====================================================
    # SHEET 2 — PERFORMANCE
    # =====================================================

    performance_sheet = workbook.create_sheet(
        "Concern Performance"
    )


    performance_headers = [

        "Concern",
        "Avg. Resolution Time",
        "Overdue",
        "Resolution Rate",

    ]


    performance_sheet.append(
        performance_headers
    )


    for cell in performance_sheet[1]:

        cell.font = Font(
            bold=True,
            color="FFFFFF"
        )

        cell.fill = PatternFill(
            fill_type="solid",
            fgColor="4F46E5"
        )

        cell.alignment = Alignment(
            horizontal="center"
        )


    for row in concerns:

        performance_sheet.append([

            row["name"] or "—",

            f'{row["average_days"]} day'
            + (
                "s"
                if row["average_days"] != 1
                else ""
            ),

            row["overdue"],

            row["resolution_rate"] / 100,

        ])


    # =====================================================
    # RESOLUTION RATE FORMAT
    # =====================================================

    for cell in performance_sheet["D"][1:]:

        cell.number_format = "0.0%"


    # =====================================================
    # COLUMN WIDTH
    # =====================================================

    performance_sheet.column_dimensions["A"].width = 35
    performance_sheet.column_dimensions["B"].width = 25
    performance_sheet.column_dimensions["C"].width = 12
    performance_sheet.column_dimensions["D"].width = 20


    # =====================================================
    # RESPONSE
    # =====================================================

    response = HttpResponse(
        content_type=(
            "application/vnd.openxmlformats-"
            "officedocument.spreadsheetml.sheet"
        )
    )


    response[
        "Content-Disposition"
    ] = 'attachment; filename="concern_report.xlsx"'


    workbook.save(response)


    return response

# ============================================================
# DASHBOARD / SUMMARY REPORT — PDF
# ============================================================

def export_reports_pdf(request):

    tickets = get_filtered_tickets(request)

    # --------------------------------------------------------
    # SUMMARY
    # --------------------------------------------------------

    total = tickets.count()

    pending = tickets.filter(
        status="pending"
    ).count()

    progress = tickets.filter(
        status="progress"
    ).count()

    resolved = tickets.filter(
        status="resolved"
    ).count()

    cancelled = tickets.filter(
        status="cancelled"
    ).count()

    overdue = tickets.filter(
        deadline__lt=timezone.now()
    ).exclude(
        status__in=["resolved", "cancelled"]
    ).count()

    resolved_on_time = tickets.filter(
        status="resolved",
        resolve_at__isnull=False,
        deadline__isnull=False,
        resolve_at__lte=F("deadline")
    ).count()

    reopened_total = (
        TicketStatusLog.objects
        .filter(
            old_status="resolved",
            ticket__in=tickets
        )
        .values("ticket")
        .distinct()
        .count()
    )

    # --------------------------------------------------------
    # DEPARTMENTS
    # --------------------------------------------------------

    departments = (
        tickets
        .values("department__name")
        .annotate(
            total=Count("id"),
            pending=Count(
                "id",
                filter=Q(status="pending")
            ),
            progress=Count(
                "id",
                filter=Q(status="progress")
            ),
            resolved=Count(
                "id",
                filter=Q(status="resolved")
            ),
            cancelled=Count(
                "id",
                filter=Q(status="cancelled")
            ),
        )
        .order_by("department__name")
    )

    # --------------------------------------------------------
    # CONCERNS
    # --------------------------------------------------------

    concerns = (
        tickets
        .values(
            "concern_type__name"
        )
        .annotate(
            total=Count("id")
        )
        .order_by("-total")
    )

    # --------------------------------------------------------
    # OUTLETS
    # --------------------------------------------------------

    outlets = (
        tickets
        .values("outlet__name")
        .annotate(
            total=Count("id"),
            pending=Count(
                "id",
                filter=Q(status="pending")
            ),
            progress=Count(
                "id",
                filter=Q(status="progress")
            ),
            resolved=Count(
                "id",
                filter=Q(status="resolved")
            ),
            cancelled=Count(
                "id",
                filter=Q(status="cancelled")
            ),
        )
        .order_by("outlet__name")
    )

    # --------------------------------------------------------
    # TECHNICIANS
    # --------------------------------------------------------

    technician_rows = []

    technicians = Technician.objects.all().order_by(
        "full_name"
    )

    selected_department = request.GET.get(
        "department",
        ""
    ).strip()

    if selected_department:

        technicians = technicians.filter(
            department__name=selected_department
        )

    for tech in technicians:

        assigned_tickets = tickets.filter(
            Q(assigned_to=tech) |
            Q(additional_technicians=tech)
        ).distinct()

        current_assigned = assigned_tickets.exclude(
            status__in=[
                "resolved",
                "cancelled"
            ]
        ).count()

        primary_total = (
            TicketAssignmentLog.objects
            .filter(
                new_technician=tech,
                ticket__in=tickets
            )
            .count()
        )

        additional_total = (
            TicketAdditionalAssignmentLog.objects
            .filter(
                technician=tech,
                action="added",
                ticket__in=tickets
            )
            .count()
        )

        total_assigned = (
            primary_total +
            additional_total
        )

        resolved_count = assigned_tickets.filter(
            status="resolved"
        ).count()

        reopened = (
            TicketStatusLog.objects
            .filter(
                technician=tech,
                old_status="resolved",
                ticket__in=tickets
            )
            .values("ticket")
            .distinct()
            .count()
        )

        technician_rows.append([
            tech.full_name or "Unnamed Technician",
            current_assigned,
            total_assigned,
            resolved_count,
            reopened,
        ])

    # ========================================================
    # PDF
    # ========================================================

    response = HttpResponse(
        content_type="application/pdf"
    )

    response["Content-Disposition"] = (
        'attachment; filename="dashboard_report.pdf"'
    )

    doc = SimpleDocTemplate(
        response,
        pagesize=landscape(A4),
        rightMargin=30,
        leftMargin=30,
        topMargin=30,
        bottomMargin=30,
    )

    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        "DashboardTitle",
        parent=styles["Title"],
        fontSize=18,
        alignment=TA_CENTER,
        spaceAfter=10,
    )

    elements = []

    elements.append(
        Paragraph(
            "FB Management — Dashboard Summary",
            title_style
        )
    )

    # --------------------------------------------------------
    # FILTER INFORMATION
    # --------------------------------------------------------

    start = request.GET.get("start", "")
    end = request.GET.get("end", "")
    department = request.GET.get("department", "")
    outlet = request.GET.get("outlet", "")

    filter_text = (
        f"Outlet: {outlet or 'All Outlets'} | "
        f"Department: {department or 'All Departments'} | "
        f"Date: {start or 'All'} to {end or 'All'}"
    )

    elements.append(
        Paragraph(
            filter_text,
            styles["Normal"]
        )
    )

    elements.append(
        Spacer(1, 15)
    )

    # --------------------------------------------------------
    # SUMMARY TABLE
    # --------------------------------------------------------

    elements.append(
        Paragraph(
            "Ticket Summary",
            styles["Heading2"]
        )
    )

    summary_data = [
        [
            "Total",
            "Pending",
            "In Progress",
            "Resolved",
            "Overdue",
            "On Time",
            "Reopened",
            "Cancelled",
        ],
        [
            total,
            pending,
            progress,
            resolved,
            overdue,
            resolved_on_time,
            reopened_total,
            cancelled,
        ],
    ]

    summary_table = Table(
        summary_data,
        repeatRows=1
    )

    summary_table.setStyle(
        TableStyle([
            (
                "BACKGROUND",
                (0, 0),
                (-1, 0),
                colors.HexColor("#2563EB")
            ),
            (
                "TEXTCOLOR",
                (0, 0),
                (-1, 0),
                colors.white
            ),
            (
                "FONTNAME",
                (0, 0),
                (-1, 0),
                "Helvetica-Bold"
            ),
            (
                "GRID",
                (0, 0),
                (-1, -1),
                0.5,
                colors.HexColor("#E5E7EB")
            ),
            (
                "ALIGN",
                (0, 0),
                (-1, -1),
                "CENTER"
            ),
            (
                "TOPPADDING",
                (0, 0),
                (-1, -1),
                8
            ),
            (
                "BOTTOMPADDING",
                (0, 0),
                (-1, -1),
                8
            ),
        ])
    )

    elements.append(summary_table)

    elements.append(
        Spacer(1, 20)
    )

    # --------------------------------------------------------
    # DEPARTMENT TABLE
    # --------------------------------------------------------

    elements.append(
        Paragraph(
            "Department Summary",
            styles["Heading2"]
        )
    )

    department_data = [[
        "Department",
        "Total",
        "Pending",
        "In Progress",
        "Resolved",
        "Cancelled",
    ]]

    for row in departments:

        department_data.append([
            row["department__name"] or "N/A",
            row["total"],
            row["pending"],
            row["progress"],
            row["resolved"],
            row["cancelled"],
        ])

    department_table = Table(
        department_data,
        repeatRows=1
    )

    department_table.setStyle(
        TableStyle([
            (
                "BACKGROUND",
                (0, 0),
                (-1, 0),
                colors.HexColor("#4F46E5")
            ),
            (
                "TEXTCOLOR",
                (0, 0),
                (-1, 0),
                colors.white
            ),
            (
                "FONTNAME",
                (0, 0),
                (-1, 0),
                "Helvetica-Bold"
            ),
            (
                "GRID",
                (0, 0),
                (-1, -1),
                0.5,
                colors.HexColor("#E5E7EB")
            ),
            (
                "FONTSIZE",
                (0, 0),
                (-1, -1),
                8
            ),
        ])
    )

    elements.append(
        department_table
    )

    elements.append(
        Spacer(1, 20)
    )

    # --------------------------------------------------------
    # OUTLET TABLE
    # --------------------------------------------------------

    elements.append(
        Paragraph(
            "Outlet Summary",
            styles["Heading2"]
        )
    )

    outlet_data = [[
        "Outlet",
        "Total",
        "Pending",
        "In Progress",
        "Resolved",
        "Cancelled",
    ]]

    for row in outlets:

        outlet_data.append([
            row["outlet__name"] or "N/A",
            row["total"],
            row["pending"],
            row["progress"],
            row["resolved"],
            row["cancelled"],
        ])

    outlet_table = Table(
        outlet_data,
        repeatRows=1
    )

    outlet_table.setStyle(
        TableStyle([
            (
                "BACKGROUND",
                (0, 0),
                (-1, 0),
                colors.HexColor("#059669")
            ),
            (
                "TEXTCOLOR",
                (0, 0),
                (-1, 0),
                colors.white
            ),
            (
                "FONTNAME",
                (0, 0),
                (-1, 0),
                "Helvetica-Bold"
            ),
            (
                "GRID",
                (0, 0),
                (-1, -1),
                0.5,
                colors.HexColor("#E5E7EB")
            ),
            (
                "FONTSIZE",
                (0, 0),
                (-1, -1),
                8
            ),
        ])
    )

    elements.append(
        outlet_table
    )

    elements.append(
        Spacer(1, 20)
    )

    # --------------------------------------------------------
    # CONCERN TABLE
    # --------------------------------------------------------

    elements.append(
        Paragraph(
            "Concern Summary",
            styles["Heading2"]
        )
    )

    concern_data = [[
        "Concern",
        "Total",
    ]]

    for row in concerns:

        concern_data.append([
            row["concern_type__name"] or "N/A",
            row["total"],
        ])

    concern_table = Table(
        concern_data,
        repeatRows=1
    )

    concern_table.setStyle(
        TableStyle([
            (
                "BACKGROUND",
                (0, 0),
                (-1, 0),
                colors.HexColor("#D97706")
            ),
            (
                "TEXTCOLOR",
                (0, 0),
                (-1, 0),
                colors.white
            ),
            (
                "FONTNAME",
                (0, 0),
                (-1, 0),
                "Helvetica-Bold"
            ),
            (
                "GRID",
                (0, 0),
                (-1, -1),
                0.5,
                colors.HexColor("#E5E7EB")
            ),
            (
                "FONTSIZE",
                (0, 0),
                (-1, -1),
                8
            ),
        ])
    )

    elements.append(
        concern_table
    )

    elements.append(
        Spacer(1, 20)
    )

    # --------------------------------------------------------
    # TECHNICIAN TABLE
    # --------------------------------------------------------

    elements.append(
        Paragraph(
            "Technician Summary",
            styles["Heading2"]
        )
    )

    technician_data = [[
        "Technician",
        "Current Assigned",
        "Total Assignments",
        "Resolved",
        "Reopened",
    ]]

    technician_data.extend(
        technician_rows
    )

    technician_table = Table(
        technician_data,
        repeatRows=1
    )

    technician_table.setStyle(
        TableStyle([
            (
                "BACKGROUND",
                (0, 0),
                (-1, 0),
                colors.HexColor("#7C3AED")
            ),
            (
                "TEXTCOLOR",
                (0, 0),
                (-1, 0),
                colors.white
            ),
            (
                "FONTNAME",
                (0, 0),
                (-1, 0),
                "Helvetica-Bold"
            ),
            (
                "GRID",
                (0, 0),
                (-1, -1),
                0.5,
                colors.HexColor("#E5E7EB")
            ),
            (
                "FONTSIZE",
                (0, 0),
                (-1, -1),
                8
            ),
        ])
    )

    elements.append(
        technician_table
    )

    doc.build(elements)

    return response

# ============================================================
# DASHBOARD / SUMMARY REPORT — EXCEL
# ============================================================

def export_reports_excel(request):

    tickets = get_filtered_tickets(request)

    # --------------------------------------------------------
    # SUMMARY
    # --------------------------------------------------------

    total = tickets.count()

    pending = tickets.filter(
        status="pending"
    ).count()

    progress = tickets.filter(
        status="progress"
    ).count()

    resolved = tickets.filter(
        status="resolved"
    ).count()

    cancelled = tickets.filter(
        status="cancelled"
    ).count()

    overdue = tickets.filter(
        deadline__lt=timezone.now()
    ).exclude(
        status__in=["resolved", "cancelled"]
    ).count()

    resolved_on_time = tickets.filter(
        status="resolved",
        resolve_at__isnull=False,
        deadline__isnull=False,
        resolve_at__lte=F("deadline")
    ).count()

    reopened_total = (
        TicketStatusLog.objects
        .filter(
            old_status="resolved",
            ticket__in=tickets
        )
        .values("ticket")
        .distinct()
        .count()
    )

    # --------------------------------------------------------
    # WORKBOOK
    # --------------------------------------------------------

    wb = Workbook()

    # ========================================================
    # SHEET 1 — SUMMARY
    # ========================================================

    ws = wb.active
    ws.title = "Summary"

    ws.append([
        "Metric",
        "Total"
    ])

    ws.append([
        "Total Tickets",
        total
    ])

    ws.append([
        "Pending",
        pending
    ])

    ws.append([
        "In Progress",
        progress
    ])

    ws.append([
        "Resolved",
        resolved
    ])

    ws.append([
        "Overdue",
        overdue
    ])

    ws.append([
        "Resolved On Time",
        resolved_on_time
    ])

    ws.append([
        "Reopened",
        reopened_total
    ])

    ws.append([
        "Cancelled",
        cancelled
    ])

    # ========================================================
    # SHEET 2 — DEPARTMENTS
    # ========================================================

    ws2 = wb.create_sheet(
        "Departments"
    )

    ws2.append([
        "Department",
        "Total",
        "Pending",
        "In Progress",
        "Resolved",
        "Cancelled",
    ])

    departments = (
        tickets
        .values("department__name")
        .annotate(
            total=Count("id"),
            pending=Count(
                "id",
                filter=Q(status="pending")
            ),
            progress=Count(
                "id",
                filter=Q(status="progress")
            ),
            resolved=Count(
                "id",
                filter=Q(status="resolved")
            ),
            cancelled=Count(
                "id",
                filter=Q(status="cancelled")
            ),
        )
        .order_by("department__name")
    )

    for row in departments:

        ws2.append([
            row["department__name"] or "N/A",
            row["total"],
            row["pending"],
            row["progress"],
            row["resolved"],
            row["cancelled"],
        ])

    # ========================================================
    # SHEET 3 — OUTLETS
    # ========================================================

    ws3 = wb.create_sheet(
        "Outlets"
    )

    ws3.append([
        "Outlet",
        "Total",
        "Pending",
        "In Progress",
        "Resolved",
        "Cancelled",
    ])

    outlets = (
        tickets
        .values("outlet__name")
        .annotate(
            total=Count("id"),
            pending=Count(
                "id",
                filter=Q(status="pending")
            ),
            progress=Count(
                "id",
                filter=Q(status="progress")
            ),
            resolved=Count(
                "id",
                filter=Q(status="resolved")
            ),
            cancelled=Count(
                "id",
                filter=Q(status="cancelled")
            ),
        )
        .order_by("outlet__name")
    )

    for row in outlets:

        ws3.append([
            row["outlet__name"] or "N/A",
            row["total"],
            row["pending"],
            row["progress"],
            row["resolved"],
            row["cancelled"],
        ])

    # ========================================================
    # SHEET 4 — CONCERNS
    # ========================================================

    ws4 = wb.create_sheet(
        "Concerns"
    )

    ws4.append([
        "Concern",
        "Total",
    ])

    concerns = (
        tickets
        .values("concern_type__name")
        .annotate(
            total=Count("id")
        )
        .order_by("-total")
    )

    for row in concerns:

        ws4.append([
            row["concern_type__name"] or "N/A",
            row["total"],
        ])

    # ========================================================
    # SHEET 5 — TECHNICIANS
    # ========================================================

    ws5 = wb.create_sheet(
        "Technicians"
    )

    ws5.append([
        "Technician",
        "Current Assigned",
        "Total Assignments",
        "Resolved",
        "Reopened",
    ])

    technicians = Technician.objects.all().order_by(
        "full_name"
    )

    selected_department = request.GET.get(
        "department",
        ""
    ).strip()

    if selected_department:

        technicians = technicians.filter(
            department__name=selected_department
        )

    for tech in technicians:

        assigned_tickets = tickets.filter(
            Q(assigned_to=tech) |
            Q(additional_technicians=tech)
        ).distinct()

        current_assigned = assigned_tickets.exclude(
            status__in=[
                "resolved",
                "cancelled"
            ]
        ).count()

        primary_total = (
            TicketAssignmentLog.objects
            .filter(
                new_technician=tech,
                ticket__in=tickets
            )
            .count()
        )

        additional_total = (
            TicketAdditionalAssignmentLog.objects
            .filter(
                technician=tech,
                action="added",
                ticket__in=tickets
            )
            .count()
        )

        total_assigned = (
            primary_total +
            additional_total
        )

        resolved_count = assigned_tickets.filter(
            status="resolved"
        ).count()

        reopened = (
            TicketStatusLog.objects
            .filter(
                technician=tech,
                old_status="resolved",
                ticket__in=tickets
            )
            .values("ticket")
            .distinct()
            .count()
        )

        ws5.append([
            tech.full_name or "Unnamed Technician",
            current_assigned,
            total_assigned,
            resolved_count,
            reopened,
        ])

    # ========================================================
    # STYLE ALL SHEETS
    # ========================================================

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="2563EB"
    )

    header_font = Font(
        bold=True,
        color="FFFFFF"
    )

    for worksheet in wb.worksheets:

        for cell in worksheet[1]:

            cell.fill = header_fill
            cell.font = header_font

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center"
            )

        worksheet.freeze_panes = "A2"

        for column_cells in worksheet.columns:

            length = 0

            column = column_cells[0].column

            for cell in column_cells:

                if cell.value:

                    length = max(
                        length,
                        len(str(cell.value))
                    )

            worksheet.column_dimensions[
                get_column_letter(column)
            ].width = min(
                length + 3,
                50
            )

    # ========================================================
    # RESPONSE
    # ========================================================

    output = BytesIO()

    wb.save(output)

    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    )

    response["Content-Disposition"] = (
        'attachment; filename="dashboard_report.xlsx"'
    )

    return response



